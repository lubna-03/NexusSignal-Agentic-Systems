import sqlite3
import pandas as pd
import os
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

# Standard Paths
DB_PATH = 'd:/NexusSignal/leads.db'
ENV_PATH = 'd:/NexusSignal/.env'
OUTPUT_FILE = 'C:/Users/HP/OneDrive/NexusSignal/NexusSignal_Master_Audit_2026.xlsx'

load_dotenv(ENV_PATH)
genai.configure(api_key=os.getenv('GEMINI_API_KEY'))

def finalize_audit():
    global OUTPUT_FILE
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    
    # print("--- INJECTING HIGH-FIDELITY FOUNDER DATA ---")
    # code removed for speed since database is already enriched
    
    print("--- GENERATING HUMANIZED AI PITCHES (GEMINI 2.0 FLASH) ---", flush=True)
    model = genai.GenerativeModel('gemini-2.0-flash')
    
    # Only fetch leads with missing pitches (including 'None' strings from previous failures)
    cursor.execute("SELECT lead_name, tech_stack FROM prospects WHERE priority='GOLD_LEAD' AND (ai_pitch IS NULL OR ai_pitch='' OR ai_pitch='None')")
    gold_leads = cursor.fetchall()
    
    import time
    import random
    for company, tech in gold_leads:
        print(f"Generating pitch for: {company}...", flush=True)
        prompt = f"""
        You are a Senior Solutions Architect reaching out to a startup founder. Peer-to-peer, zero sales fluff.
        Company: {company}
        Current Tech Context: {tech}
        
        STRICT RULES:
        - BANNED WORDS: 'Unlock', 'Leverage', 'Cutting-edge', 'I hope this finds you well', 'Transform', 'Empower'.
        - Tone: Helpful, low-key, professional.
        
        Structure (3 sentences EXACTLY):
        1. The Hook: A technical observation (e.g. checkout feels heavy/load lag).
        2. The Value: Explain the cost (mobile conversion/scaling pain).
        3. The CTA: Soft question about a migration roadmap.
        
        Output ONLY the 3 sentences.
        """
        success = False
        models_to_try = ['gemini-2.0-flash', 'gemini-pro-latest']
        
        for model_name in models_to_try:
            if success: break
            print(f"Attempting model: {model_name}", flush=True)
            current_model = genai.GenerativeModel(model_name)
            try:
                # Set a short timeout for quick failover
                response = current_model.generate_content(prompt)
                pitch = response.text.strip()
                if pitch and len(pitch) > 10:
                    cursor.execute("UPDATE prospects SET ai_pitch=? WHERE lead_name=?", (pitch, company))
                    conn.commit()
                    print("DONE: Pitch Saved.", flush=True)
                    success = True
                    time.sleep(2)
            except Exception as e:
                print(f"Quota issue with {model_name}.", flush=True)

        if not success:
            print(f"Using Template Fallback for {company}...", flush=True)
            hooks = [
                f"I noticed some significant latency in {company}'s mobile checkout flow recently.",
                f"Your current tech stack seems to be hitting some scaling bottlenecks during peak traffic.",
                f"The architecture at {company} looks like it could benefit from some modernization to reduce technical debt."
            ]
            values = [
                "This typically results in a 15-20% drop in mobile conversion rates.",
                "Left unmanaged, this technical debt will likely stall your next feature release cycle.",
                "High infrastructure costs are often a side effect of these legacy bottlenecks."
            ]
            ctas = [
                "Do you have a migration roadmap in place for the current quarter?",
                "Are you open to a brief technical audit of the checkout path?",
                "Who on your team is currently leading the infrastructure scale-up?"
            ]
            
            pitch = f"{random.choice(hooks)} {random.choice(values)} {random.choice(ctas)}"
            cursor.execute("UPDATE prospects SET ai_pitch=? WHERE lead_name=?", (pitch, company))
            conn.commit()
            print("DONE: Template Pitch Saved.", flush=True)
    
    print("--- Pitch Generation Cycle Complete ---", flush=True)
    
    print("--- CONSOLIDATING FINAL REPORT ---", flush=True)
    df = pd.read_sql_query("SELECT priority, tech_status, lead_name, website_url, tech_stack, contact_name, contact_email, contact_phone, ai_pitch FROM prospects", conn)
    
    # Readability Map
    lead_type_map = {'MODERNIZATION': 'Legacy Tech', 'NO_WEBSITE': 'No Website', 'BROKEN': 'Broken Site'}
    df['tech_status'] = df['tech_status'].map(lambda x: lead_type_map.get(x, x))
    
    # Outreach Tier Logic
    df['Outreach Tier'] = df['contact_phone'].apply(lambda x: 'P1: Direct Dial' if (x and str(x).strip() and str(x).lower() != 'none') else 'P2: Email Only')
    
    # Column Renaming
    mapping = {
        'priority': 'Tier/Status',
        'tech_status': 'Lead Type',
        'lead_name': 'Company',
        'website_url': 'Website',
        'tech_stack': 'Technical Debt',
        'contact_name': 'Decision Maker',
        'contact_email': 'Email',
        'contact_phone': 'Phone',
        'ai_pitch': 'AI Generated Pitch'
    }
    df = df.rename(columns=mapping)
    df = df.sort_values(by='Tier/Status', ascending=False)
    
    counter = 1
    final_output = OUTPUT_FILE
    while True:
        try:
            df.to_excel(final_output, index=False)
            print(f"File Saved: {final_output}")
            break
        except PermissionError:
            counter += 1
            final_output = OUTPUT_FILE.replace('.xlsx', f'_v{counter}.xlsx')
            if counter > 10: 
                print("Too many locked files. Please close Excel.")
                return None
    
    # Update OUTPUT_FILE for subsequent formatting
    OUTPUT_FILE = final_output

    # Formatting
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid') # Gold for Elite
    yellow_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid') # Lemon Chiffon for Verified
    blue_fill = PatternFill(start_color='E1F5FE', end_color='E1F5FE', fill_type='solid') # Light Sky Blue for Standard
    
    # Highlighting
    status_col = None
    email_col = None
    pitch_col_letter = None
    for i, cell in enumerate(ws[1]):
        if cell.value == 'Tier/Status': status_col = i + 1
        if cell.value == 'Email': email_col = i + 1
        if cell.value == 'AI Generated Pitch': pitch_col_letter = cell.column_letter

    for row in ws.iter_rows(min_row=2):
        is_gold = row[status_col-1].value == 'GOLD_LEAD'
        has_email = row[email_col-1].value and len(str(row[email_col-1].value)) > 5
        
        # Color Logic: Gold -> Elite, Yellow -> Verified, Blue -> Standard
        if is_gold:
            fill_to_use = gold_fill
        elif has_email:
            fill_to_use = yellow_fill
        else:
            fill_to_use = blue_fill
        
        for cell in row:
            cell.fill = fill_to_use
            if is_gold:
                cell.font = Font(bold=True) # Bold for Gold leads
    
    # Widths & Wrapping
    for col in ws.columns:
        column = col[0].column_letter
        if column == pitch_col_letter:
            ws.column_dimensions[column].width = 60
            for cell in col:
                cell.alignment = Alignment(wrap_text=False)
        else:
            ws.column_dimensions[column].width = 20

    wb.save(OUTPUT_FILE)
    conn.close()
    print("Formatting Complete.")
    return OUTPUT_FILE

if __name__ == "__main__":
    finalize_audit()
